from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from database import DB_PATH, connect, init_schema

BASE_DIR = Path(__file__).resolve().parent
IMPORT_DIR = BASE_DIR / "resources" / "imports"


def read_rows(path: Path) -> list[tuple]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.active
    return list(worksheet.iter_rows(values_only=True))


def clear_database() -> None:
    if DB_PATH.exists():
        DB_PATH.unlink()


def seed_reference_tables(connection) -> None:
    roles = ["Гость", "Авторизированный клиент", "Менеджер", "Администратор"]
    categories = set()
    manufacturers = set()
    suppliers = set()
    units = set()

    product_rows = read_rows(IMPORT_DIR / "Tovar.xlsx")[1:]
    for row in product_rows:
        if row[0] is None:
            continue
        categories.add(str(row[6] or "Без категории").strip())
        manufacturers.add(str(row[5] or "Не указан").strip())
        suppliers.add(str(row[4] or "Не указан").strip())
        units.add(str(row[2] or "шт.").strip())

    for table_name, values in {
        "role": roles,
        "category": sorted(categories),
        "manufacturer": sorted(manufacturers),
        "supplier": sorted(suppliers),
        "unit": sorted(units),
    }.items():
        for value in values:
            connection.execute(f"INSERT INTO {table_name}(name) VALUES (?)", (value,))

    pickup_rows = read_rows(IMPORT_DIR / "Пункты выдачи_import.xlsx")[1:]
    for row in pickup_rows:
        if row and row[0]:
            connection.execute("INSERT INTO pickup_point(address) VALUES (?)", (str(row[0]).strip(),))

    for status_name in ["Новый", "Завершен"]:
        connection.execute("INSERT INTO order_status(name) VALUES (?)", (status_name,))


def lookup_id(connection, table: str, name: str) -> int:
    row = connection.execute(f"SELECT id FROM {table} WHERE name = ?", (name,)).fetchone()
    return int(row["id"])


def seed_users(connection) -> dict[str, int]:
    users: dict[str, int] = {}
    user_rows = read_rows(IMPORT_DIR / "user_import.xlsx")[1:]
    for row in user_rows:
        if row[0] is None:
            continue
        role_name = str(row[0]).strip() if row[0] else "Авторизированный клиент"
        role_id = lookup_id(connection, "role", role_name)
        connection.execute(
            "INSERT INTO user(full_name, login, password, role_id) VALUES (?, ?, ?, ?)",
            (str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip(), role_id),
        )
        user_id = int(connection.execute("SELECT last_insert_rowid()").fetchone()[0])
        users[str(row[1]).strip()] = user_id
    return users


def seed_products(connection) -> None:
    product_rows = read_rows(IMPORT_DIR / "Tovar.xlsx")[1:]
    for row in product_rows:
        if row[0] is None:
            continue
        image_name = str(row[10]).strip() if row[10] else "picture.png"
        image_path = f"resources/images/{image_name}"
        connection.execute(
            """
            INSERT INTO product(
                article, name, unit_id, price, supplier_id, manufacturer_id, category_id,
                discount_percent, stock_quantity, description, image_path
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                str(row[0]).strip(),
                str(row[1]).strip(),
                lookup_id(connection, "unit", str(row[2] or "шт.").strip()),
                float(row[3] or 0),
                lookup_id(connection, "supplier", str(row[4] or "Не указан").strip()),
                lookup_id(connection, "manufacturer", str(row[5] or "Не указан").strip()),
                lookup_id(connection, "category", str(row[6] or "Без категории").strip()),
                int(row[7] or 0),
                int(row[8] or 0),
                str(row[9] or "").strip(),
                image_path,
            ),
        )


def parse_order_items(raw_value: str) -> list[tuple[str, int]]:
    items = [part.strip() for part in str(raw_value).split(",") if str(part).strip()]
    result: list[tuple[str, int]] = []
    for index in range(0, len(items), 2):
        article = items[index]
        quantity = int(items[index + 1])
        result.append((article, quantity))
    return result


def seed_orders(connection, users: dict[str, int]) -> None:
    order_rows = read_rows(IMPORT_DIR / "Заказ_import.xlsx")[1:]
    for row in order_rows:
        if row[0] is None:
            continue
        connection.execute(
            """
            INSERT INTO customer_order(
                id, order_date, delivery_date, pickup_point_id, customer_id, receive_code, status_id
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(row[0]),
                row[2].strftime("%Y-%m-%d") if hasattr(row[2], "strftime") else str(row[2]),
                row[3].strftime("%Y-%m-%d") if hasattr(row[3], "strftime") else str(row[3]),
                int(row[4]),
                users[str(row[5]).strip()],
                int(row[6]),
                lookup_id(connection, "order_status", str(row[7]).strip()),
            ),
        )
        for article, quantity in parse_order_items(str(row[1])):
            connection.execute(
                "INSERT INTO order_item(order_id, product_article, quantity) VALUES (?, ?, ?)",
                (int(row[0]), article, quantity),
            )


def main() -> None:
    clear_database()
    init_schema()
    with connect() as connection:
        seed_reference_tables(connection)
        users = seed_users(connection)
        seed_products(connection)
        seed_orders(connection, users)
        connection.commit()
    print(f"База данных создана: {DB_PATH}")


if __name__ == "__main__":
    main()
